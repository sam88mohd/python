#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Feb 28 21:16:20 2023

@author: sm
"""

import openpyxl, pprint

def count_tract(fname):
    
        wb = openpyxl.load_workbook(fname)
        ws = wb.active
        
        count_tract_dict = {}
        
        for row in range(2, ws.max_row + 1):
            pop = ws[f"D{row}"].value
            state = ws[f"B{row}"].value
            country = ws[f"C{row}"].value
            
            count_tract_dict.setdefault(state, {})
            count_tract_dict[state].setdefault(country, {'tract': 0, 'pop': 0})
            count_tract_dict[state][country]['tract'] += 1
            count_tract_dict[state][country]['pop'] += int(pop)
            
        return count_tract_dict
    
        
def write_to_file(file, content):
    try:
        with open(file, 'w') as f:
            f.write(f"alldata = {pprint.pformat(content)}")
    except Exception as err:
        print(err)
        
if __name__ == '__main__':
    import sys
    fname = sys.argv[1]
    new_file = sys.argv[2]
    write_to_file(new_file, count_tract(fname))