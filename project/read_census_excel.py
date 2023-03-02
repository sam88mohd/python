#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 27 22:33:32 2023

@author: sm
"""

import openpyxl, pprint

def read_census_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        country_data = {}
        
        for row in range(2, ws.max_row + 1):
            state = ws[f"B{row}"].value
            country = ws[f"C{row}"].value
            populations = ws[f"D{row}"].value
            
            country_data.setdefault(state, {})
            country_data[state].setdefault(country, {'tract': 0, 'pop': 0}) 
            country_data[state][country]['tract'] += 1
            country_data[state][country]['pop'] += int(populations)
    
        return country_data
    except FileNotFoundError as err:
        print(err)

def write_to_file(file, line):
    try:
        with open(file, 'w') as f:
            f.write('alldata = ' + pprint.pformat(line))
    except Exception as err:
        print(err)

if __name__ == '__main__':
    import sys
    filename = sys.argv[1]
    file = sys.argv[2]
    write_to_file(file, read_census_excel(filename))